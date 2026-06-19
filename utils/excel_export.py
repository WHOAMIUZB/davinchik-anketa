# -*- coding: utf-8 -*-
"""Foydalanuvchilar ma'lumotlarini Excel (.xlsx) fayliga eksport qilish."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

import database as db
from config import EXCEL_PATH


async def export_users_excel(path: str = EXCEL_PATH) -> str:
    users = await db.get_users_for_excel()

    wb = Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"

    headers = ["№", "Ism", "Yosh", "Jinsi", "Viloyat", "O'zi haqida tarif", "Holat", "Telegram ID"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    status_labels = {
        "pending": "Kutilmoqda",
        "approved": "Tasdiqlangan",
        "rejected": "Rad etilgan",
        "none": "Anketa yo'q",
    }

    for i, u in enumerate(users, start=1):
        ws.append([
            i,
            u.get("name"),
            u.get("age"),
            u.get("gender"),
            u.get("region"),
            u.get("bio"),
            status_labels.get(u.get("status"), u.get("status")),
            u.get("telegram_id"),
        ])

    widths = [5, 20, 8, 10, 18, 45, 15, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    wb.save(path)
    return path
